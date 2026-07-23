from __future__ import annotations

import base64
from datetime import date, datetime, timedelta
import html
import json
import os
from pathlib import Path
import re
import sys
import time
import wave

import requests
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"
SITE = ROOT / "site"
LOG_DIR = ROOT / "logs"
XLSX = OUT / "2026_21天禁禱禱告指引_繁體中文.xlsx"

MODEL = os.environ.get("GEMINI_TTS_MODEL", "gemini-2.5-flash-preview-tts")
VOICE = os.environ.get("GEMINI_TTS_VOICE", "Laomedeia")
WAIT_SECONDS = float(os.environ.get("GEMINI_TTS_WAIT_SECONDS", "8"))
PAUSE_SECONDS = float(os.environ.get("DAY_AUDIO_PAUSE_SECONDS", "0.45"))
SAMPLE_RATE = 24000
CHANNELS = 1
SAMPLE_WIDTH = 2
START_DATE = date(2026, 7, 13)
TOTAL_DAYS = 21


def get_secret(name: str) -> str | None:
    """Read a secret without logging it, including newly-set Windows user env vars."""
    value = os.environ.get(name)
    if value or os.name != "nt":
        return value
    try:
        import winreg

        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Environment") as key:
            value, _ = winreg.QueryValueEx(key, name)
            return str(value) if value else None
    except (FileNotFoundError, OSError):
        return None


def log(event: str, message: str, **extra: object) -> None:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    entry = {
        "time": datetime.now().isoformat(timespec="seconds"),
        "event": event,
        "message": message,
        **extra,
    }
    with (LOG_DIR / "automation_events.jsonl").open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


def normalize_text(s: str) -> str:
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"(?<=[\u4e00-\u9fff])[ \t]+(?=[\u4e00-\u9fff])", "", s)
    s = re.sub(r"\s+([，。；：！？、）」』])", r"\1", s)
    s = re.sub(r"([「『（])\s+", r"\1", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def split_paragraphs(content: str) -> tuple[str, str, list[str]]:
    content = normalize_text(content)
    lines = [ln.strip() for ln in content.splitlines() if ln.strip()]
    day_label = lines[0]
    topic = lines[1] if len(lines) > 1 else ""
    body_lines = lines[2:]
    paragraphs: list[str] = []
    cur: list[str] = []
    for ln in body_lines:
        if ln.startswith("請與我們一起禱告"):
            if cur:
                paragraphs.append("".join(cur).strip())
                cur = []
            paragraphs.append(ln)
            continue
        cur.append(ln)
        joined = "".join(cur)
        if re.search(r"[。！？）」』]$", joined) and len(joined) >= 35:
            paragraphs.append(joined.strip())
            cur = []
    if cur:
        paragraphs.append("".join(cur).strip())

    final: list[str] = []
    for p in paragraphs:
        if p.startswith("主啊，") and len(p) > 180:
            final.extend([x for x in re.split(r"(?<=[。！？])", p) if x.strip()])
        else:
            final.append(p)
    return day_label, topic, final


def tts_text(s: str) -> str:
    s = re.sub(r"（[^）]*）", "", s)
    s = re.sub(r"\([^)]*\)", "", s)
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"(?<=[\u4e00-\u9fff])\s+(?=[\u4e00-\u9fff])", "", s)
    s = re.sub(r"\s+([，。；：！？、）」』])", r"\1", s)
    return s.strip()


def load_day(day_index: int) -> dict[str, object]:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not (1 <= day_index <= len(rows)):
        raise ValueError(f"day_index out of range: {day_index}")
    row = rows[day_index - 1]
    day_label, topic, paragraphs = split_paragraphs(str(row[2]))
    return {
        "day_index": day_index,
        "day_label": day_label,
        "date": str(row[1]),
        "topic": topic,
        "paragraphs": paragraphs,
    }


def make_chunks(day: dict[str, object]) -> list[str]:
    items = [
        str(day["day_label"]),
        str(day["date"]),
        str(day["topic"]),
        *[tts_text(str(p)) for p in day["paragraphs"]],
    ]
    chunks: list[str] = []
    cur = ""
    for item in [x for x in items if x.strip()]:
        if len(cur) + len(item) > 520 and cur:
            chunks.append(cur.strip())
            cur = ""
        cur += ("\n\n" if cur else "") + item
    if cur:
        chunks.append(cur.strip())
    return chunks


CSS = """
:root{--bg:#f6efe3;--card:#fffdf8;--ink:#243044;--muted:#687083;--accent:#9b6a2f;--accent2:#fff1d7;--line:rgba(93,63,30,.16)}
*{box-sizing:border-box}body{margin:0;font-family:-apple-system,BlinkMacSystemFont,"Noto Sans TC","Microsoft JhengHei",sans-serif;background:radial-gradient(circle at top left,rgba(255,223,162,.65),transparent 28rem),linear-gradient(180deg,#fbf5e8,#f4efe8);color:var(--ink);line-height:1.95}main{width:min(100%,760px);margin:0 auto;padding:22px 15px 48px}.card{background:rgba(255,253,248,.96);border:1px solid var(--line);border-radius:26px;box-shadow:0 18px 46px rgba(72,47,20,.13);overflow:hidden}header{padding:28px 22px 22px;background:linear-gradient(145deg,rgba(151,100,42,.14),rgba(255,255,255,0));border-bottom:1px solid var(--line)}.meta{display:flex;gap:10px;flex-wrap:wrap;color:#6d461d;font-weight:800}.pill{padding:5px 12px;border-radius:999px;background:var(--accent2);border:1px solid rgba(155,106,47,.2);font-size:14px}h1{margin:16px 0 4px;font-size:clamp(31px,8vw,44px);line-height:1.18}.subtitle{margin:0;color:var(--muted);font-size:15px}.player{margin-top:22px;padding:14px;border-radius:18px;background:#fff8eb;border:1px solid rgba(155,106,47,.2)}.player-label{margin:0 0 9px;color:#6d461d;font-weight:800;font-size:14px}audio{width:100%;display:block;accent-color:var(--accent)}section.content{padding:22px}.paragraph,.scripture,.prayer-heading{font-size:18px;margin:0 0 18px;letter-spacing:.01em}.scripture{padding:16px 17px;border-radius:18px;background:#f8f1e6;border-left:5px solid var(--accent);color:#384052}.prayer-heading{margin-top:24px;color:#805322;font-weight:900;font-size:20px}.footer-note{margin-top:28px;padding-top:18px;border-top:1px solid var(--line);color:var(--muted);font-size:14px}@media(max-width:430px){main{padding:12px 10px 34px}.card{border-radius:22px}header{padding:24px 18px 20px}section.content{padding:20px 17px}.paragraph,.scripture,.prayer-heading{font-size:17px;line-height:1.98}}
""".strip()


def render_day_html(day: dict[str, object]) -> str:
    day_index = int(day["day_index"])
    audio_name = f"day-{day_index:02d}.wav"
    para_html = []
    for p in day["paragraphs"]:
        p = str(p)
        cls = "prayer-heading" if p.startswith("請與我們一起禱告") else "scripture" if p.startswith("「") else "paragraph"
        para_html.append(f'<p class="{cls}">{html.escape(p)}</p>')
    return f"""<!doctype html>
<html lang="zh-Hant-TW">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(str(day["day_label"]))}｜{html.escape(str(day["topic"]))}</title>
  <style>{CSS}</style>
</head>
<body>
<main>
  <article class="card">
    <header>
      <div class="meta"><span class="pill">{html.escape(str(day["day_label"]))}</span><span class="pill">{html.escape(str(day["date"]))}</span></div>
      <h1>{html.escape(str(day["topic"]))}</h1>
      <p class="subtitle">2026 21天禁食禱告指引｜手機閱讀版</p>
      <div class="player">
        <p class="player-label">朗讀音檔</p>
        <audio controls preload="metadata" src="{audio_name}"></audio>
      </div>
    </header>
    <section class="content">
      {chr(10).join(para_html)}
      <p class="footer-note">朗讀版已移除括號內經文出處；閱讀版保留原文經文參照。</p>
    </section>
  </article>
</main>
</body>
</html>
"""


def render_index(days: list[dict[str, object]]) -> str:
    links = []
    for day in days:
        i = int(day["day_index"])
        links.append(
            f'<a class="day" href="day-{i:02d}.html">{html.escape(str(day["day_label"]))}｜{html.escape(str(day["date"]))}｜{html.escape(str(day["topic"]))}</a>'
        )
    return f"""<!doctype html>
<html lang="zh-Hant-TW"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>2026 21天禁禱禱告指引</title><style>{CSS} .day-list{{padding:22px}} a.day{{display:block;text-decoration:none;color:var(--ink);background:#fff8eb;border:1px solid var(--line);border-radius:18px;padding:16px 18px;margin:12px 0;font-weight:800}}</style></head><body><main><article class="card"><header><div class="meta"><span class="pill">2026</span><span class="pill">21天禁禱</span></div><h1>2026 21天禁禱禱告指引</h1><p class="subtitle">手機閱讀與朗讀音檔</p></header><section class="day-list">{''.join(links)}</section></article></main></body></html>"""


def render_index(days: list[dict[str, object]]) -> str:
    """Render the index with the canonical 2026 programme title."""
    links = []
    for day in days:
        i = int(day["day_index"])
        links.append(
            f'<a class="day" href="day-{i:02d}.html">{html.escape(str(day["day_label"]))}｜{html.escape(str(day["date"]))}｜{html.escape(str(day["topic"]))}</a>'
        )
    return f'''<!doctype html>
<html lang="zh-Hant-TW"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>2026 21天禁食禱告指引</title><style>{CSS} .day-list{{padding:22px}} a.day{{display:block;text-decoration:none;color:var(--ink);background:#fff8eb;border:1px solid var(--line);border-radius:18px;padding:16px 18px;margin:12px 0;font-weight:800}}</style></head>
<body><main><article class="card"><header><div class="meta"><span class="pill">2026</span><span class="pill">21天禁食禱告</span></div><h1>2026 21天禁食禱告指引</h1><p class="subtitle">手機閱讀與朗讀音檔</p></header><section class="day-list">{''.join(links)}</section></article></main></body></html>'''


def find_audio_data(payload: object) -> str | None:
    if isinstance(payload, dict):
        output_audio = payload.get("output_audio")
        if isinstance(output_audio, dict) and isinstance(output_audio.get("data"), str):
            return output_audio["data"]
        for key in ("data", "audio"):
            if isinstance(payload.get(key), str):
                return payload[key]
        for value in payload.values():
            found = find_audio_data(value)
            if found:
                return found
    elif isinstance(payload, list):
        for item in payload:
            found = find_audio_data(item)
            if found:
                return found
    return None


def write_wav(path: Path, pcm: bytes) -> None:
    with wave.open(str(path), "wb") as wf:
        wf.setnchannels(CHANNELS)
        wf.setsampwidth(SAMPLE_WIDTH)
        wf.setframerate(SAMPLE_RATE)
        wf.writeframes(pcm)


def read_wav_pcm(path: Path) -> bytes:
    with wave.open(str(path), "rb") as wf:
        return wf.readframes(wf.getnframes())


def request_tts(api_key: str, chunk: str, index: int, total: int) -> bytes:
    prompt = (
        "Read the following Traditional Chinese devotional text exactly as written. "
        "Use a warm, gentle, lively Taiwanese Mandarin female voice. "
        "Keep the pace steady and light; do not slow down near the end. "
        "Sound natural, clear, peaceful, and encouraging. "
        "Do not read bracketed scripture references, because they have already been removed. "
        "Pause only briefly between sentences.\n\n"
        "[warm, gentle, lively, steady pace, Taiwanese Mandarin]\n\n"
        f"{chunk}"
    )
    response = requests.post(
        "https://generativelanguage.googleapis.com/v1beta/interactions",
        headers={"x-goog-api-key": api_key, "Content-Type": "application/json"},
        json={
            "model": MODEL,
            "input": prompt,
            "response_format": {"type": "audio"},
            "generation_config": {"speech_config": [{"voice": VOICE}]},
        },
        timeout=240,
    )
    if response.status_code >= 400:
        err = LOG_DIR / f"tts_error_chunk_{index:02d}.json"
        err.write_text(response.text, encoding="utf-8")
        raise RuntimeError(f"Gemini TTS failed for chunk {index}/{total}; see {err}")
    audio_b64 = find_audio_data(response.json())
    if not audio_b64:
        raise RuntimeError(f"No audio returned for chunk {index}/{total}")
    return base64.b64decode(audio_b64)


def generate_audio(day_index: int, chunks: list[str]) -> Path:
    api_key = get_secret("GEMINI_API_KEY") or get_secret("GOOGLE_API_KEY")
    if not api_key:
        raise RuntimeError("Missing GEMINI_API_KEY")
    chunk_paths = []
    for i, chunk in enumerate(chunks, 1):
        log("tts_chunk_started", f"開始產生第 {day_index} 天音檔第 {i}/{len(chunks)} 段。", day=day_index, chunk=i)
        pcm = request_tts(api_key, chunk, i, len(chunks))
        p = OUT / f"day-{day_index:02d}-chunk-{i:02d}.wav"
        write_wav(p, pcm)
        chunk_paths.append(p)
        log("tts_chunk_completed", f"完成第 {day_index} 天音檔第 {i}/{len(chunks)} 段。", day=day_index, chunk=i, file=str(p))
        if i < len(chunks):
            time.sleep(WAIT_SECONDS)
    pause = b"\x00" * int(SAMPLE_RATE * PAUSE_SECONDS) * SAMPLE_WIDTH * CHANNELS
    merged = bytearray()
    for i, p in enumerate(chunk_paths):
        if i:
            merged.extend(pause)
        merged.extend(read_wav_pcm(p))
    audio = SITE / f"day-{day_index:02d}.wav"
    write_wav(audio, bytes(merged))
    return audio


def determine_day_index() -> int:
    if len(sys.argv) > 1:
        return int(sys.argv[1])
    target = date.today() + timedelta(days=1)
    return (target - START_DATE).days + 1


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    SITE.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    day_index = determine_day_index()
    if not 1 <= day_index <= TOTAL_DAYS:
        target = date.today() + timedelta(days=1)
        log(
            "pipeline_skipped",
            f"目標日期 {target.isoformat()} 不在 21 天禁禱期間內，略過製作。",
            target_date=target.isoformat(),
            calculated_day=day_index,
        )
        print(json.dumps({"skipped": True, "target_date": target.isoformat()}, ensure_ascii=False))
        return
    log("pipeline_started", f"開始製作第 {day_index} 天。", day=day_index)
    day = load_day(day_index)
    chunks = make_chunks(day)
    (OUT / f"day-{day_index:02d}-tts-script.txt").write_text("\n\n---CHUNK---\n\n".join(chunks), encoding="utf-8")
    (SITE / f"day-{day_index:02d}.html").write_text(render_day_html(day), encoding="utf-8")
    # The local generator keeps prior pages in site/, whereas the cloud
    # workflow checks out previously published pages at the repository root.
    # Include both locations so rebuilding the homepage never drops history.
    available_indexes: set[int] = set()
    for base in (SITE, ROOT):
        for p in base.glob("day-*.html"):
            m = re.search(r"day-(\d+)\.html$", p.name)
            if m:
                available_indexes.add(int(m.group(1)))
    available_days = [load_day(i) for i in sorted(available_indexes)]
    (SITE / "index.html").write_text(render_index(available_days), encoding="utf-8")
    log("html_completed", f"完成第 {day_index} 天 HTML 與首頁更新。", day=day_index, file=str(SITE / f"day-{day_index:02d}.html"))
    audio = generate_audio(day_index, chunks)
    log("audio_completed", f"完成第 {day_index} 天音檔。", day=day_index, file=str(audio))
    log("pipeline_completed", f"第 {day_index} 天本機製作完成，等待上傳 GitHub Pages。", day=day_index)
    print(json.dumps({"day": day_index, "html": str(SITE / f"day-{day_index:02d}.html"), "audio": str(audio)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
